import openai
from openpyxl import load_workbook

from config import *

# 设置 OpenAI 的 API Key 和 API 基础地址
openai.api_key = api_key
openai.base_url = base_url
openai.default_headers = {"x-foo": "true"}

# 读取 Excel 文件
file_path = 'test-data.xlsx'
wb = load_workbook(file_path)
ws = wb.active  # 获取活动的工作表

# 获取标题行
headers = [cell.value for cell in ws[1]]
type_index = headers.index("题型")
content_index = headers.index("题目内容")
answer_index = headers.index("参考答案")

# 获取选项的索引
a_index = headers.index("a") if "a" in headers else None
b_index = headers.index("b") if "b" in headers else None
c_index = headers.index("c") if "c" in headers else None
d_index = headers.index("d") if "d" in headers else None
e_index = headers.index("e") if "e" in headers else None

# 设置P列的表头为"解析"
ws['P1'] = "解析"

# 逐行处理 Excel 文件中的数据
for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True),
                          start=2):  # 从第二行开始（假设第一行是标题）

    # 检查是否所有字段都为空
    if all(cell is None for cell in row):
        break

    question_type = str(row[type_index])
    question_content = str(row[content_index])
    reference_answer = str(row[answer_index])

    # 拼接选项内容
    options = []
    if a_index is not None and row[a_index] is not None:
        options.append(f"a: {row[a_index]}")
    if b_index is not None and row[b_index] is not None:
        options.append(f"b: {row[b_index]}")
    if c_index is not None and row[c_index] is not None:
        options.append(f"c: {row[c_index]}")
    if d_index is not None and row[d_index] is not None:
        options.append(f"d: {row[d_index]}")
    if e_index is not None and row[e_index] is not None:
        options.append(f"e: {row[e_index]}")
    options_text = '; '.join(options)

    # 将标题和对应的内容拼接成字符串
    input_text = f"题型: {question_type}; 题目内容: {question_content}; {options_text}; 正确答案: {reference_answer}"

    # 使用 OpenAI 进行处理
    completion = openai.chat.completions.create(
        model=model_name,
        messages=[
            {
                "role": "user",
                "content": input_text,
            },
        ],
    )

    # 获取生成的回复内容
    output = completion.choices[0].message.content

    # 写入Excel的P列
    ws[f'P{idx}'].value = output

    print(f"Processed row {idx}")

# 保存修改后的Excel文件
wb.save('result.xlsx')
