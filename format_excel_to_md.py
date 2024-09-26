from openpyxl import load_workbook

# 读取 Excel 文件
file_path = 'result.xlsx'
wb = load_workbook(file_path)
ws = wb.active  # 获取活动的工作表

# 定义列索引
headers = ["试题编码", "考察能力", "知识点", "试题来源", "题型", "难度", "分数", "时长（秒）",
           "题目内容", "参考答案", "a", "b", "c", "d", "e", "解析"]
column_indices = {header: headers.index(header) for header in headers}

# 打开输出文件
with open('formatted.md', 'w', encoding='utf-8') as output_file:
    # 逐行处理 Excel 文件中的数据
    for row in ws.iter_rows(min_row=2, values_only=True):  # 从第二行开始（假设第一行是标题）
        # 检查是否所有字段都为空
        if all(cell is None for cell in row):
            break

        question_code = str(row[column_indices["试题编码"]])
        question_type = str(row[column_indices["题型"]])
        question_content = str(row[column_indices["题目内容"]])
        reference_answer = str(row[column_indices["参考答案"]])
        explanation = str(row[column_indices["解析"]])

        # 格式化输出
        output_file.write(f"### {question_code}\n\n")
        output_file.write(f"- [ ] **题目**：{question_content}\n")

        output_file.write(f"({question_type})\n\n")

        # 只有在非判断题的情况下才输出选项
        if "判断" not in question_type:
            output_file.write("**选项**：\n")
            for option in ['a', 'b', 'c', 'd', 'e']:
                if row[column_indices[option]] is not None:
                    output_file.write(f"{option}: {row[column_indices[option]]}\n")
            output_file.write("\n")

        # 格式化正确答案，确保占据5个字符的空间
        formatted_answer = reference_answer.ljust(5)
        output_file.write("**答案**：\n")
        output_file.write(f"~~=={formatted_answer}==~~\n\n")

        output_file.write("**解析**：\n\n")
        output_file.write(f"{explanation}\n\n---\n\n")

print("格式化完成，请查看 formatted.md 文件。")